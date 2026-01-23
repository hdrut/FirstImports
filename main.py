from __future__ import annotations

import os
import re

from datetime import datetime, timezone

APP_STARTED_AT_UTC = datetime.now(timezone.utc)

def get_build_info() -> dict:
    """Return build/deploy metadata for display in templates.

    On Koyeb GitHub deployments, KOYEB_GIT_SHA is provided automatically.
    """
    sha = os.getenv("KOYEB_GIT_SHA") or os.getenv("GIT_SHA") or ""
    branch = os.getenv("KOYEB_GIT_BRANCH") or os.getenv("GIT_BRANCH") or ""
    # Optional: let you override the displayed build time from an env var (ISO 8601).
    build_time = os.getenv("APP_BUILD_TIME") or os.getenv("BUILD_TIME") or ""

    if not build_time:
        build_time = APP_STARTED_AT_UTC.isoformat()

    return {
        "sha": sha,
        "sha_short": sha[:7] if sha else "",
        "branch": branch,
        "build_time": build_time,  # ISO string (usually UTC)
    }


from datetime import datetime, date
from typing import Optional, List

from fastapi import FastAPI, Request, Form, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, Response, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from starlette.middleware.sessions import SessionMiddleware

from passlib.context import CryptContext

from sqlalchemy import (
    create_engine, String, Integer, Float, Boolean, Date, DateTime, ForeignKey, select, func, desc, text
)
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, relationship, Session, sessionmaker

from openpyxl import Workbook
from openpyxl.utils import get_column_letter



# -----------------------------
# Config
# -----------------------------
APP_NAME = "Ventas First"
SECRET_KEY = os.getenv("APP_SECRET", "CAMBIAME-POR-UNA-CLAVE-LARGA")
DB_URL = os.getenv("DATABASE_URL", "sqlite:///./ventas.db")

# Render suele dar postgres://... (SQLAlchemy prefiere postgresql+psycopg2://...)
if DB_URL.startswith("postgres://"):
    DB_URL = DB_URL.replace("postgres://", "postgresql+psycopg2://", 1)

DEFAULT_ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "admin@local")
DEFAULT_ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

engine = create_engine(DB_URL, connect_args={"check_same_thread": False} if DB_URL.startswith("sqlite") else {})
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)


class Base(DeclarativeBase):
    pass


# -----------------------------
# Models
# -----------------------------
class User(Base):
    __tablename__ = "users"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    nombre: Mapped[str] = mapped_column(String(60))
    apellido: Mapped[str] = mapped_column(String(60))
    email: Mapped[str] = mapped_column(String(200), unique=True, index=True)
    password_hash: Mapped[str] = mapped_column(String(255))
    rol: Mapped[str] = mapped_column(String(20), default="user")  # "admin" | "user"
    activo: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)

    ventas: Mapped[List["Venta"]] = relationship(back_populates="creado_por_user")


class Articulo(Base):
    __tablename__ = "articulos"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    nombre: Mapped[str] = mapped_column(String(160), index=True)
    categoria: Mapped[str] = mapped_column(String(60), default="Otros")
    activo: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)

    items: Mapped[List["VentaItem"]] = relationship(back_populates="articulo")


class Venta(Base):
    __tablename__ = "ventas"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    numero: Mapped[str] = mapped_column(String(30), unique=True, index=True)  # V-YYYY-000123
    fecha: Mapped[date] = mapped_column(Date, default=date.today)

    cliente_nombre: Mapped[str] = mapped_column(String(80))
    cliente_apellido: Mapped[str] = mapped_column(String(80))
    telefono: Mapped[Optional[str]] = mapped_column(String(40), nullable=True)
    direccion: Mapped[Optional[str]] = mapped_column(String(200), nullable=True)

    medio_pago: Mapped[Optional[str]] = mapped_column(String(30), nullable=True)  # Efectivo/Transferencia/Tarjeta/Mixto
    observaciones: Mapped[Optional[str]] = mapped_column(String(500), nullable=True)

    total: Mapped[float] = mapped_column(Float, default=0.0)

    creado_por: Mapped[int] = mapped_column(ForeignKey("users.id"))
    creado_por_user: Mapped[User] = relationship(back_populates="ventas")

    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)

    items: Mapped[List["VentaItem"]] = relationship(back_populates="venta", cascade="all, delete-orphan")


class VentaItem(Base):
    __tablename__ = "venta_items"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    venta_id: Mapped[int] = mapped_column(ForeignKey("ventas.id"), index=True)
    articulo_id: Mapped[Optional[int]] = mapped_column(ForeignKey("articulos.id"), nullable=True)

    descripcion_libre: Mapped[Optional[str]] = mapped_column(String(200), nullable=True)
    cantidad: Mapped[float] = mapped_column(Float, default=1.0)
    precio_unitario: Mapped[float] = mapped_column(Float, default=0.0)
    subtotal: Mapped[float] = mapped_column(Float, default=0.0)

    venta: Mapped[Venta] = relationship(back_populates="items")
    articulo: Mapped[Optional[Articulo]] = relationship(back_populates="items")


# -----------------------------
# App init
# -----------------------------
app = FastAPI(title=APP_NAME)
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, same_site="lax")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


def db() -> Session:
    d = SessionLocal()
    try:
        yield d
    finally:
        d.close()


# -----------------------------
# Password helpers (bcrypt-safe)
# -----------------------------
def _bcrypt_safe_pw(pw: str) -> str:
    # bcrypt admite máximo 72 BYTES (no caracteres)
    b = pw.encode("utf-8")
    if len(b) <= 72:
        return pw
    return b[:72].decode("utf-8", errors="ignore")


def hash_pw(pw: str) -> str:
    return pwd_context.hash(_bcrypt_safe_pw(pw))


def verify_pw(pw: str, ph: str) -> bool:
    return pwd_context.verify(_bcrypt_safe_pw(pw), ph)


# -----------------------------
# DB init + lightweight migrations
# -----------------------------
def migrate():
    # Agrega columna direccion si falta (Postgres / SQLite compatibles con IF NOT EXISTS?)
    # Postgres: OK. SQLite: ADD COLUMN funciona, pero sin IF NOT EXISTS en versiones viejas.
    # Como hoy corrés en Postgres, usamos IF NOT EXISTS.
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE ventas ADD COLUMN IF NOT EXISTS direccion VARCHAR(200);"))


def init_db():
    # Primero: crear tablas
    Base.metadata.create_all(bind=engine)

    # Migraciones livianas
    try:
        migrate()
    except Exception:
        # Si falla (p.ej. SQLite viejo), seguimos sin romper el arranque.
        pass

    # Crear admin si no existe
    with SessionLocal() as s:
        exists = s.scalar(select(func.count()).select_from(User))
        if not exists:
            admin = User(
                nombre="Admin",
                apellido="Sistema",
                email=DEFAULT_ADMIN_EMAIL,
                password_hash=hash_pw(DEFAULT_ADMIN_PASSWORD),
                rol="admin",
                activo=True,
            )
            s.add(admin)
            s.commit()


init_db()


# -----------------------------
# Auth helpers
# -----------------------------
def get_current_user(request: Request, d: Session = Depends(db)) -> User:
    uid = request.session.get("uid")
    if not uid:
        raise HTTPException(status_code=401)
    user = d.get(User, uid)
    if not user or not user.activo:
        raise HTTPException(status_code=401)
    return user


def require_admin(user: User = Depends(get_current_user)) -> User:
    if user.rol != "admin":
        raise HTTPException(status_code=403)
    return user


def redirect(url: str) -> RedirectResponse:
    return RedirectResponse(url=url, status_code=303)


# -----------------------------
# Utilities
# -----------------------------
def next_sale_number(d: Session, y: int) -> str:
    prefix = f"V-{y}-"
    last = d.scalar(
        select(Venta.numero)
        .where(Venta.numero.like(f"{prefix}%"))
        .order_by(desc(Venta.id))
        .limit(1)
    )
    if not last:
        return f"{prefix}000001"
    m = re.match(rf"V-{y}-(\d+)$", last)
    n = int(m.group(1)) if m else 0
    return f"{prefix}{n + 1:06d}"


def parse_date(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def money(x: float) -> str:
    # Formato AR
    return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def ws_autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


# -----------------------------
# Routes: Auth
# -----------------------------
@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    uid = request.session.get("uid")
    return redirect("/dashboard") if uid else redirect("/login")


@app.get("/login", response_class=HTMLResponse)
def login_get(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "app": APP_NAME})


@app.post("/login")
def login_post(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    d: Session = Depends(db),
):
    user = d.scalar(select(User).where(User.email == email.lower().strip()))
    if not user or not user.activo or not verify_pw(password, user.password_hash):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "app": APP_NAME, "error": "Credenciales inválidas."},
            status_code=400,
        )
    request.session["uid"] = user.id
    return redirect("/dashboard")


@app.post("/logout")
def logout(request: Request):
    request.session.clear()
    return redirect("/login")


# -----------------------------
# Routes: Dashboard
# -----------------------------
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    d: Session = Depends(db),
    user: User = Depends(get_current_user),
):
    df = parse_date(date_from) or date(2025, 1, 1) #date.today().replace(day=1)
    dt = parse_date(date_to) or date.today()

    total = d.scalar(select(func.coalesce(func.sum(Venta.total), 0.0)).where(Venta.fecha.between(df, dt))) or 0.0
    count = d.scalar(select(func.count()).select_from(Venta).where(Venta.fecha.between(df, dt))) or 0
    ticket = (total / count) if count else 0.0

    top_art = d.execute(
        select(
            Articulo.nombre,
            func.coalesce(func.sum(VentaItem.subtotal), 0.0).label("fact"),
            func.coalesce(func.sum(VentaItem.cantidad), 0.0).label("qty"),
        )
        .join(VentaItem, VentaItem.articulo_id == Articulo.id)
        .join(Venta, Venta.id == VentaItem.venta_id)
        .where(Venta.fecha.between(df, dt))
        .group_by(Articulo.nombre)
        .order_by(desc("fact"))
        .limit(10)
    ).all()

    ult = d.execute(
        select(Venta)
        .where(Venta.fecha.between(df, dt))
        .order_by(desc(Venta.fecha), desc(Venta.id))
        .limit(10)
    ).scalars().all()

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "app": APP_NAME,
            "user": user,
            "df": df,
            "dt": dt,
            "total": total,
            "count": count,
            "ticket": ticket,
            "top_art": top_art,
            "ult": ult,
            "money": money,
            "build": get_build_info(),
        },
    )


# -----------------------------
# Routes: Ventas
# -----------------------------
@app.get("/ventas", response_class=HTMLResponse)
def ventas_list(
    request: Request,
    q: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vendedor: Optional[int] = None,
    d: Session = Depends(db),
    user: User = Depends(get_current_user),
):
    df = parse_date(date_from)
    dt = parse_date(date_to)

    stmt = select(Venta).order_by(desc(Venta.fecha), desc(Venta.id))

    if df and dt:
        stmt = stmt.where(Venta.fecha.between(df, dt))
    elif df:
        stmt = stmt.where(Venta.fecha >= df)
    elif dt:
        stmt = stmt.where(Venta.fecha <= dt)

    if vendedor:
        stmt = stmt.where(Venta.creado_por == vendedor)

    if q and q.strip():
        s = f"%{q.strip().lower()}%"
        stmt = stmt.where(
            func.lower(Venta.cliente_nombre).like(s)
            | func.lower(Venta.cliente_apellido).like(s)
            | func.lower(Venta.numero).like(s)
            | func.lower(func.coalesce(Venta.telefono, "")).like(s)
            | func.lower(func.coalesce(Venta.direccion, "")).like(s)
        )

    ventas = d.execute(stmt.limit(500)).scalars().all()
    vendedores = d.execute(select(User).where(User.activo == True).order_by(User.apellido, User.nombre)).scalars().all()

    return templates.TemplateResponse(
        "sales_list.html",
        {
            "request": request,
            "app": APP_NAME,
            "user": user,
            "ventas": ventas,
            "vendedores": vendedores,
            "money": money,
            "q": q or "",
            "date_from": date_from or "",
            "date_to": date_to or "",
            "vendedor": vendedor or "",
        },
    )


@app.get("/ventas/nueva", response_class=HTMLResponse)
def venta_new_get(
    request: Request,
    d: Session = Depends(db),
    user: User = Depends(get_current_user),
):
    articulos = d.execute(select(Articulo).where(Articulo.activo == True).order_by(Articulo.categoria, Articulo.nombre)).scalars().all()
    medios = ["Efectivo", "Transferencia", "Tarjeta", "Mixto"]
    return templates.TemplateResponse(
        "sale_form.html",
        {
            "request": request,
            "app": APP_NAME,
            "user": user,
            "modo": "new",
            "venta": None,
            "items": [],
            "articulos": articulos,
            "medios": medios,
            "today": date.today().isoformat(),
        },
    )


@app.post("/ventas/nueva")
def venta_new_post(
    request: Request,
    fecha: str = Form(...),
    cliente_nombre: str = Form(...),
    cliente_apellido: str = Form(...),
    telefono: str = Form(""),
    direccion: str = Form(""),
    medio_pago: str = Form(""),
    observaciones: str = Form(""),
    item_articulo_id: List[str] = Form([]),
    item_desc: List[str] = Form([]),
    item_qty: List[str] = Form([]),
    item_unit: List[str] = Form([]),
    d: Session = Depends(db),
    user: User = Depends(get_current_user),
):
    f = parse_date(fecha)
    if not f:
        raise HTTPException(400, "Fecha inválida")
    y = f.year
    numero = next_sale_number(d, y)

    venta = Venta(
        numero=numero,
        fecha=f,
        cliente_nombre=cliente_nombre.strip(),
        cliente_apellido=cliente_apellido.strip(),
        telefono=telefono.strip() or None,
        direccion=direccion.strip() or None,
        medio_pago=medio_pago.strip() or None,
        observaciones=observaciones.strip() or None,
        creado_por=user.id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
        total=0.0,
    )

    total = 0.0
    for aid, qty_txt, unit_txt in zip(item_articulo_id, item_qty, item_unit):
        if (aid or "").strip() == "":
            continue
    
        try:
            qty = float((qty_txt or "0").replace(",", "."))
            unit = float((unit_txt or "0").replace(",", "."))
        except ValueError:
            qty, unit = 0.0, 0.0
    
        subtotal = qty * unit
        total += subtotal
    
        item = VentaItem(
            articulo_id=int(aid) if (aid or "").strip() else None,
            descripcion_libre=(desc_txt or "").strip(),
            cantidad=qty,
            precio_unitario=unit,
            subtotal=subtotal,
        )
        venta.items.append(item)

    venta.total = total
    d.add(venta)
    d.commit()
    return redirect(f"/ventas/{venta.id}")


@app.get("/ventas/{venta_id}", response_class=HTMLResponse)
def venta_view(
    request: Request,
    venta_id: int,
    d: Session = Depends(db),
    user: User = Depends(get_current_user),
):
    venta = d.get(Venta, venta_id)
    if not venta:
        raise HTTPException(404)
    items = d.execute(select(VentaItem).where(VentaItem.venta_id == venta_id)).scalars().all()
    creador = d.get(User, venta.creado_por)
    return templates.TemplateResponse(
        "sale_view.html",
        {"request": request, "app": APP_NAME, "user": user, "venta": venta, "items": items, "creador": creador, "money": money},
    )


@app.get("/ventas/{venta_id}/editar", response_class=HTMLResponse)
def venta_edit_get(
    request: Request,
    venta_id: int,
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    venta = d.get(Venta, venta_id)
    if not venta:
        raise HTTPException(404)
    items = d.execute(select(VentaItem).where(VentaItem.venta_id == venta_id)).scalars().all()
    articulos = d.execute(select(Articulo).where(Articulo.activo == True).order_by(Articulo.categoria, Articulo.nombre)).scalars().all()
    medios = ["Efectivo", "Transferencia", "Tarjeta", "Mixto"]
    return templates.TemplateResponse(
        "sale_form.html",
        {
            "request": request,
            "app": APP_NAME,
            "user": user,
            "modo": "edit",
            "venta": venta,
            "items": items,
            "articulos": articulos,
            "medios": medios,
            "today": date.today().isoformat(),
        },
    )


from itertools import zip_longest

@app.post("/ventas/{venta_id}/editar")
def venta_edit_post(
    request: Request,
    venta_id: int,
    fecha: str = Form(...),
    cliente_nombre: str = Form(...),
    cliente_apellido: str = Form(...),
    telefono: str = Form(""),
    direccion: str = Form(""),
    medio_pago: str = Form(""),
    observaciones: str = Form(""),
    item_articulo_id: List[str] = Form([]),
    item_desc: List[str] = Form([]),
    item_qty: List[str] = Form([]),
    item_unit: List[str] = Form([]),
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    venta = d.get(Venta, venta_id)
    if not venta:
        raise HTTPException(404)

    f = parse_date(fecha)
    if not f:
        raise HTTPException(400, "Fecha inválida")

    # --- Actualizar cabecera ---
    venta.fecha = f
    venta.cliente_nombre = cliente_nombre.strip()
    venta.cliente_apellido = cliente_apellido.strip()
    venta.telefono = telefono.strip() or None
    venta.direccion = direccion.strip() or None
    venta.medio_pago = medio_pago.strip() or None
    venta.observaciones = observaciones.strip() or None

    # --- Helpers ---
    def parse_num(txt: str) -> float:
        s = (txt or "").strip().replace(" ", "")
        if not s:
            return 0.0

        # Si tiene coma, asumimos formato AR: 1.234,56
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            # Si no tiene coma, permitimos decimal con punto: 1234.56
            s = s.replace(",", "")

        try:
            return float(s)
        except ValueError:
            return 0.0

    # --- Reemplazar items ---
    venta.items.clear()   # ✅ solo una vez

    total = 0.0

    # ✅ zip_longest evita que se “corte” todo si item_desc viene vacío
    for aid, desc_txt, qty_txt, unit_txt in zip_longest(
        item_articulo_id, item_desc, item_qty, item_unit, fillvalue=""
    ):
        aid = (aid or "").strip()
        desc_txt = (desc_txt or "").strip()

        # Si no hay artículo ni desc, ignorar
        if aid == "" and desc_txt == "":
            continue

        qty = parse_num(qty_txt)



@app.post("/ventas/{venta_id}/eliminar")
def venta_delete(
    venta_id: int,
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    venta = d.get(Venta, venta_id)
    if not venta:
        raise HTTPException(404)
    d.delete(venta)
    d.commit()
    return redirect("/ventas")


# -----------------------------
# Routes: Artículos
# -----------------------------
@app.get("/articulos", response_class=HTMLResponse)
def articulos_list(
    request: Request,
    d: Session = Depends(db),
    user: User = Depends(get_current_user),
):
    articulos = (
        d.execute(select(Articulo).order_by(Articulo.activo.desc(), Articulo.categoria, Articulo.nombre))
        .scalars()
        .all()
    )
    return templates.TemplateResponse(
        "articles_list.html",
        {"request": request, "app": APP_NAME, "user": user, "articulos": articulos},
    )


@app.post("/articulos/nuevo")
def articulo_new(
    nombre: str = Form(...),
    categoria: str = Form("Otros"),
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    a = Articulo(nombre=nombre.strip(), categoria=categoria.strip() or "Otros", activo=True)
    d.add(a)
    d.commit()
    return redirect("/articulos")


@app.post("/articulos/{articulo_id}/toggle")
def articulo_toggle(
    articulo_id: int,
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    a = d.get(Articulo, articulo_id)
    if not a:
        raise HTTPException(404)
    a.activo = not a.activo
    d.commit()
    return redirect("/articulos")


@app.get("/articulos/{articulo_id}/editar", response_class=HTMLResponse)
def articulo_edit(
    request: Request,
    articulo_id: int,
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    a = d.get(Articulo, articulo_id)
    if not a:
        raise HTTPException(404)
    return templates.TemplateResponse(
        "article_form.html",
        {"request": request, "app": APP_NAME, "user": user, "articulo": a},
    )


@app.post("/articulos/{articulo_id}/editar")
def articulo_edit_post(
    articulo_id: int,
    nombre: str = Form(...),
    categoria: str = Form(""),
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    a = d.get(Articulo, articulo_id)
    if not a:
        raise HTTPException(404)
    a.nombre = nombre.strip()
    a.categoria = categoria.strip() or "Otros"
    d.commit()
    return redirect("/articulos")
# -----------------------------
# Routes: Usuarios (admin)
# -----------------------------
@app.get("/usuarios", response_class=HTMLResponse)
def users_list(
    request: Request,
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    users = d.execute(select(User).order_by(User.activo.desc(), User.apellido, User.nombre)).scalars().all()
    return templates.TemplateResponse(
        "users_list.html",
        {"request": request, "app": APP_NAME, "user": user, "users": users},
    )


@app.post("/usuarios/nuevo")
def user_new(
    nombre: str = Form(...),
    apellido: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    rol: str = Form("user"),
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    em = email.lower().strip()
    if d.scalar(select(func.count()).select_from(User).where(User.email == em)):
        raise HTTPException(400, "Email ya existe")
    u = User(
        nombre=nombre.strip(),
        apellido=apellido.strip(),
        email=em,
        password_hash=hash_pw(password),
        rol="admin" if rol == "admin" else "user",
        activo=True,
    )
    d.add(u)
    d.commit()
    return redirect("/usuarios")


@app.post("/usuarios/{user_id}/toggle")
def user_toggle(
    user_id: int,
    d: Session = Depends(db),
    user: User = Depends(require_admin),
):
    u = d.get(User, user_id)
    if not u:
        raise HTTPException(404)
    if u.id == user.id:
        raise HTTPException(400, "No podés desactivarte a vos mismo.")
    u.activo = not u.activo
    d.commit()
    return redirect("/usuarios")


# -----------------------------
# Excel Exports
# -----------------------------
@app.get("/export/ventas.xlsx")
def export_ventas(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    d: Session = Depends(db),
    user: User = Depends(get_current_user),
):
    df = parse_date(date_from)
    dt = parse_date(date_to)

    stmt = select(Venta).order_by(desc(Venta.fecha), desc(Venta.id))
    if df and dt:
        stmt = stmt.where(Venta.fecha.between(df, dt))
    elif df:
        stmt = stmt.where(Venta.fecha >= df)
    elif dt:
        stmt = stmt.where(Venta.fecha <= dt)

    ventas = d.execute(stmt).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Número", "Fecha", "Cliente", "Teléfono", "Dirección", "Medio de pago", "Total", "Vendedor", "Observaciones"])

    for v in ventas:
        u = d.get(User, v.creado_por)
        cliente = f"{v.cliente_apellido}, {v.cliente_nombre}"
        vendedor = f"{u.apellido}, {u.nombre}" if u else ""
        ws.append([v.numero, v.fecha.isoformat(), cliente, v.telefono or "", v.direccion or "", v.medio_pago or "", v.total, vendedor, v.observaciones or ""])

    ws_autofit(ws)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ventas.xlsx"'},
    )


@app.get("/export/items.xlsx")
def export_items(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    d: Session = Depends(db),
    user: User = Depends(get_current_user),
):
    df = parse_date(date_from)
    dt = parse_date(date_to)

    stmt = (
        select(Venta, VentaItem, Articulo)
        .join(VentaItem, VentaItem.venta_id == Venta.id)
        .outerjoin(Articulo, Articulo.id == VentaItem.articulo_id)
        .order_by(desc(Venta.fecha), desc(Venta.id))
    )
    if df and dt:
        stmt = stmt.where(Venta.fecha.between(df, dt))
    elif df:
        stmt = stmt.where(Venta.fecha >= df)
    elif dt:
        stmt = stmt.where(Venta.fecha <= dt)

    rows = d.execute(stmt).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    ws.append(["Venta Nº", "Fecha", "Cliente", "Artículo", "Categoría", "Descripción", "Cantidad", "P. Unitario", "Subtotal"])

    for v, it, a in rows:
        cliente = f"{v.cliente_apellido}, {v.cliente_nombre}"
        art = a.nombre if a else ""
        cat = a.categoria if a else ""
        ws.append([v.numero, v.fecha.isoformat(), cliente, art, cat, it.descripcion_libre or "", it.cantidad, it.precio_unitario, it.subtotal])

    ws_autofit(ws)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="items.xlsx"'},
    )


@app.get("/export/resumen.xlsx")
def export_resumen(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    d: Session = Depends(db),
    user: User = Depends(get_current_user),
):
    df = parse_date(date_from)
    dt = parse_date(date_to)

    stmt = (
        select(
            Articulo.nombre,
            Articulo.categoria,
            func.coalesce(func.sum(VentaItem.cantidad), 0.0).label("cantidad"),
            func.coalesce(func.sum(VentaItem.subtotal), 0.0).label("facturacion"),
        )
        .join(VentaItem, VentaItem.articulo_id == Articulo.id)
        .join(Venta, Venta.id == VentaItem.venta_id)
        .group_by(Articulo.nombre, Articulo.categoria)
        .order_by(desc("facturacion"))
    )
    if df and dt:
        stmt = stmt.where(Venta.fecha.between(df, dt))
    elif df:
        stmt = stmt.where(Venta.fecha >= df)
    elif dt:
        stmt = stmt.where(Venta.fecha <= dt)

    rows = d.execute(stmt).all()
    total = sum(r.facturacion for r in rows) if rows else 0.0

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen por artículo"
    ws.append(["Artículo", "Categoría", "Cantidad", "Facturación", "% del total"])

    for r in rows:
        pct = (r.facturacion / total) if total else 0.0
        ws.append([r.nombre, r.categoria, r.cantidad, r.facturacion, pct])

    ws_autofit(ws)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="resumen_por_articulo.xlsx"'},
    )


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return FileResponse("static/favicon.ico")
